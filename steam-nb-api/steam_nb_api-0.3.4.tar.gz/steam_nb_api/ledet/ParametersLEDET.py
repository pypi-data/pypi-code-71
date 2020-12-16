import os
import numpy as np
import datetime
import csv
import openpyxl
from openpyxl.utils import get_column_letter
import xlsxwriter
from dataclasses import dataclass, asdict

from steam_nb_api.resources.ResourceReader import ResourceReader

@dataclass
class LEDETInputs:
    BlankRows: np.ndarray = np.array([[3,4,6,34,35,39,43,47,50,53,55,61,75,77],[3,4,6,34,35,39,43,47,50,53,55,61,75,77]])
    T00: float= 0.0
    l_magnet: float= 0.0
    I00: float= 0.0
    M_m: np.ndarray = np.array([])
    fL_I: np.ndarray = np.array([])
    fL_L: np.ndarray = np.array([])
    GroupToCoilSection: np.ndarray = np.array([])
    polarities_inGroup: np.ndarray = np.array([])
    nT: np.ndarray = np.array([])
    nStrands_inGroup: np.ndarray = np.array([])
    l_mag_inGroup: np.ndarray = np.array([])
    ds_inGroup: np.ndarray = np.array([])
    f_SC_strand_inGroup: np.ndarray = np.array([])
    f_ro_eff_inGroup: np.ndarray = np.array([])
    Lp_f_inGroup: np.ndarray = np.array([])
    RRR_Cu_inGroup: np.ndarray = np.array([])
    SCtype_inGroup: np.ndarray = np.array([])
    STtype_inGroup: np.ndarray = np.array([])
    insulationType_inGroup: np.ndarray = np.array([])
    internalVoidsType_inGroup: np.ndarray = np.array([])
    externalVoidsType_inGroup: np.ndarray = np.array([])
    wBare_inGroup: np.ndarray = np.array([])
    hBare_inGroup: np.ndarray = np.array([])
    wIns_inGroup: np.ndarray = np.array([])
    hIns_inGroup: np.ndarray = np.array([])
    Lp_s_inGroup: np.ndarray = np.array([])
    R_c_inGroup: np.ndarray = np.array([])
    Tc0_NbTi_ht_inGroup: np.ndarray = np.array([])
    Bc2_NbTi_ht_inGroup: np.ndarray = np.array([])
    c1_Ic_NbTi_inGroup: np.ndarray = np.array([])
    c2_Ic_NbTi_inGroup: np.ndarray = np.array([])
    Tc0_Nb3Sn_inGroup: np.ndarray = np.array([])
    Bc2_Nb3Sn_inGroup: np.ndarray = np.array([])
    Jc_Nb3Sn0_inGroup: np.ndarray = np.array([])
    overwrite_f_internalVoids_inGroup: np.ndarray = np.array([])
    overwrite_f_externalVoids_inGroup: np.ndarray = np.array([])
    el_order_half_turns: np.ndarray = np.array([])
    alphasDEG: np.ndarray = np.array([])
    rotation_block: np.ndarray = np.array([])
    mirror_block: np.ndarray = np.array([])
    mirrorY_block: np.ndarray = np.array([])
    iContactAlongWidth_From: np.ndarray = np.array([])
    iContactAlongWidth_To: np.ndarray = np.array([])
    iContactAlongHeight_From: np.ndarray = np.array([])
    iContactAlongHeight_To: np.ndarray = np.array([])
    iStartQuench: np.ndarray = np.array([])
    tStartQuench: np.ndarray = np.array([])
    lengthHotSpot_iStartQuench: np.ndarray = np.array([])
    vQ_iStartQuench: np.ndarray = np.array([])
    R_circuit: float= 0.0
    R_crowbar: float= 0.0
    Ud_crowbar: float= 0.0
    t_PC: float= 0.0
    t_PC_LUT: np.ndarray = np.array([])
    I_PC_LUT: np.ndarray = np.array([])
    tEE: float= 0.0
    R_EE_triggered: float= 0.0
    tCLIQ: np.ndarray = np.array([])
    directionCurrentCLIQ: np.ndarray = np.array([])
    nCLIQ: np.ndarray = np.array([])
    U0: np.ndarray = np.array([])
    C: np.ndarray = np.array([])
    Rcapa: np.ndarray = np.array([])
    tQH: np.ndarray = np.array([])
    U0_QH: np.ndarray = np.array([])
    C_QH: np.ndarray = np.array([])
    R_warm_QH: np.ndarray = np.array([])
    w_QH: np.ndarray = np.array([])
    h_QH: np.ndarray = np.array([])
    s_ins_QH: np.ndarray = np.array([])
    type_ins_QH: np.ndarray = np.array([])
    s_ins_QH_He: np.ndarray = np.array([])
    type_ins_QH_He: np.ndarray = np.array([])
    l_QH: np.ndarray = np.array([])
    f_QH: np.ndarray = np.array([])
    iQH_toHalfTurn_From: np.ndarray = np.array([])
    iQH_toHalfTurn_To: np.ndarray = np.array([])
    tQuench: np.ndarray = np.array([])
    initialQuenchTemp: np.ndarray = np.array([])
    HalfTurnToInductanceBlock: np.ndarray = np.array([])
    M_InductanceBlock_m: np.ndarray= np.array([])
@dataclass
class LEDETOptions:
    BlankRows: np.ndarray= np.array([[],[]])
    time_vector_params: np.ndarray = np.array([])
    Iref: float = 0.0
    flagIron: float = 0.0
    flagSelfField: float = 0.0
    headerLines: float = 0.0
    columnsXY: np.ndarray = np.array([])
    columnsBxBy: np.ndarray = np.array([])
    flagPlotMTF: float = 0.0
    flag_calculateInductanceMatrix: float = 0.0
    flag_useExternalInitialization: float = 0.0
    flag_initializeVar: float = 0.0
    flag_fastMode: float = 0.0
    flag_controlCurrent: float = 0.0
    flag_automaticRefinedTimeStepping: float = 0.0
    flag_IronSaturation: float = 0.0
    flag_InvertCurrentsAndFields: float = 0.0
    flag_ScaleDownSuperposedMagneticField: float = 0.0
    flag_HeCooling: float = 0.0
    fScaling_Pex: float = 0.0
    fScaling_Pex_AlongHeight: float = 0.0
    fScaling_MR: float = 0.0
    flag_scaleCoilResistance_StrandTwistPitch: float = 0.0
    flag_separateInsulationHeatCapacity: float = 0.0
    flag_ISCL: float = 0.0
    fScaling_Mif: float = 0.0
    fScaling_Mis: float = 0.0
    flag_StopIFCCsAfterQuench: float = 0.0
    flag_StopISCCsAfterQuench: float = 0.0
    tau_increaseRif: float = 0.0
    tau_increaseRis: float = 0.0
    fScaling_RhoSS: float = 0.0
    maxVoltagePC: float = 0.0
    flag_symmetricGroundingEE: float = 0.0
    flag_removeUc: float = 0.0
    BtX_background: float = 0.0
    BtY_background: float = 0.0
    flag_showFigures: float = 0.0
    flag_saveFigures: float = 0.0
    flag_saveMatFile: float = 0.0
    flag_saveTxtFiles: float = 0.0
    flag_generateReport: float = 0.0
    flag_hotSpotTemperatureInEachGroup: float = 0.0
    MinMaxXY_MTF: np.ndarray = np.array([])
@dataclass
class LEDETPlots:
    BlankRows: np.ndarray = np.array([[],[]])
    suffixPlot: str = ''
    typePlot: int = 0
    outputPlotSubfolderPlot: str = ''
    variableToPlotPlot: np.ndarray = np.array([])
    selectedStrandsPlot: np.ndarray = np.array([])
    selectedTimesPlot: np.ndarray = np.array([])
    labelColorBarPlot: np.ndarray = np.array([])
    minColorBarPlot: float = 0.0
    maxColorBarPlot: float = 0.0
    MinMaxXYPlot: np.ndarray = np.array([])
    flagSavePlot: int = 0
    flagColorPlot: int = 0
    flagInvisiblePlot: int = 0
@dataclass
class LEDETVariables:
    BlankRows: np.ndarray = np.array([[],[]])
    variableToSaveTxt: np.ndarray = np.array([])
    typeVariableToSaveTxt: np.ndarray = np.array([])
    variableToInitialize: np.ndarray = np.array([])
@dataclass
class Cable:
    A_CableInsulated: np.ndarray = np.array([])
    f_SC: np.ndarray = np.array([])
    f_ST: np.ndarray = np.array([])
    SCtype: np.ndarray = np.array([])
    STtype: np.ndarray = np.array([])
    Tc0_NbTi: np.ndarray = np.array([])
    Bc20_NbTi: np.ndarray = np.array([])
    c1_Ic_NbTi: np.ndarray = np.array([])
    c2_Ic_NbTi: np.ndarray = np.array([])
    alpha_NbTi: float = 0.59
    Jc_Nb3Sn0: np.ndarray = np.array([])
    Tc0_Nb3Sn: np.ndarray = np.array([])
    Bc20_Nb3Sn: np.ndarray = np.array([])

def read_row(workSheet, Nrow, St = False):
    rowValues = np.array([])
    row = workSheet[Nrow]
    for cell in row:
        if not St:
            if isinstance(cell.value, str): continue
        rowValues = np.append(rowValues, cell.value)
    rowValues = rowValues[rowValues != None]
    return rowValues

def CompareLEDETParameters(FileA, FileB, Precision = 1E-5):
    Diff = 0

    P_a = ParametersLEDET()
    P_a.readLEDETExcel(FileA)
    P_b = ParametersLEDET()
    P_b.readLEDETExcel(FileB)

    print("Starting Comparison of A: ({}) and B: ({})".format(FileA, FileB))

    ## Check Inputs
    for attribute in P_a.Inputs.__annotations__:
        Block = 1
        CC = 0

        P_a_a = P_a.getAttribute("Inputs", attribute)
        P_b_a = P_b.getAttribute("Inputs", attribute)

        if isinstance(P_a_a, float) or isinstance(P_a_a, int):
            if P_a_a != P_b_a:
                if Block: print("Found difference in Parameter {}, A: {}, B: {}".format(attribute,P_a_a, P_b_a))
                Block = 0
                Diff = 1
        elif len(P_a_a) != len(P_b_a):
            Diff = 1
            if Block:
                Block = 0
                print('Parameter {} of A, {} has not the same length as Parameter of B, {}'.format(attribute, len(P_a_a), len(P_b_a)))
        else:
            Pos = []
            for k in range(len(P_a_a)):
                try:
                    if P_a_a[k] != P_b_a[k]:
                        if abs(P_a_a[k]-P_b_a[k])>Precision:
                            Diff = 1
                            if Block:
                                print("Found difference in Parameter {}".format(attribute))
                                Block = 0
                            Pos.append(k)
                except:
                    for j in range(P_a_a.shape[1]):
                        if P_a_a[k,j] != P_b_a[k,j]:
                            if abs(P_a_a[k,j] - P_b_a[k,j]) > Precision:
                                Diff = 1
                                if Block:
                                    print("Found difference in Parameter {}".format(attribute))
                                    Block = 0
                                Pos.append([k,j])

            if len(Pos)>0:
                if len(Pos)<10:
                    print("Different Positions: {}".format(Pos))
                else:
                    print("Many values are different (>10)")


    ## Check Options
    for attribute in P_a.Options.__annotations__:
        Block = 1
        P_a_a = P_a.getAttribute("Options", attribute)
        P_b_a = P_b.getAttribute("Options", attribute)
        if isinstance(P_a_a, float) or isinstance(P_a_a, int):
            if P_a_a != P_b_a:
                if Block: print("Found difference in Parameter {}, A: {}, B: {}".format(attribute, P_a_a, P_b_a))
                Block = 0
                Diff = 1
        elif len(P_a_a) != len(P_b_a):
            Diff = 1
            if Block:
                Block = 0
                print(
                    'Parameter {} of A, {} has not the same length as Parameter of B, {}'.format(attribute, len(P_a_a),
                                                                                                 len(P_b_a)))
        else:
            Pos = []
            for k in range(len(P_a_a)):
                try:
                    if P_a_a[k] != P_b_a[k]:
                        if abs(P_a_a[k] - P_b_a[k]) > Precision:
                            Diff = 1
                            if Block:
                                print("Found difference in Parameter {}".format(attribute))
                                Block = 0
                            Pos.append(k)
                except:
                    for j in range(P_a_a.shape[1]):
                        if P_a_a[k, j] != P_b_a[k, j]:
                            if abs(P_a_a[k, j] - P_b_a[k, j]) > Precision:
                                Diff = 1
                                if Block:
                                    print("Found difference in Parameter {}".format(attribute))
                                    Block = 0
                                Pos.append([k, j])

            if len(Pos) > 0:
                if len(Pos) < 10:
                    print("Different Positions: {}".format(Pos))
                else:
                    print("Many values are different (>10)")

    if Diff==0:
        print("Files are equal.")

class ParametersLEDET:
    '''
        Class of LEDET parameters
    '''
    def setAttribute(self, LEDETclass, attribute, value):
        try:
            setattr(LEDETclass, attribute, value)
        except:
            setattr(getattr(self, LEDETclass), attribute, value)

    def getAttribute(self, LEDETclass, attribute):
        try:
            return getattr(LEDETclass, attribute)
        except:
            return getattr(getattr(self, LEDETclass), attribute)

    def updateBlankrows(self,LEDETclass):
        attributeCounter = 1
        lineAdd = np.zeros(LEDETclass.BlankRows[1].shape)
        if lineAdd.size == 0:
            return

        lengthAdd = 1
        currentStatus = 0
        for attribute in LEDETclass.__annotations__:
            if attribute == 'BlankRows':
                continue
            downLength = 1
            if(type(self.getAttribute(LEDETclass,attribute))==np.ndarray):
                if self.getAttribute(LEDETclass, attribute).ndim > 1:
                    downLength = self.getAttribute(LEDETclass, attribute).shape[0]
            lengthAdd = lengthAdd + downLength
            if attributeCounter in LEDETclass.BlankRows[0]:
                lineAdd[currentStatus:] = lineAdd[currentStatus:]+lengthAdd
                lengthAdd = 1
                currentStatus = currentStatus +1

            attributeCounter = attributeCounter + 1

        LEDETclass.BlankRows[1] = lineAdd-1

        if len(LEDETclass.overwrite_f_internalVoids_inGroup) != 0:
            LEDETclass.BlankRows[1, 3:] = LEDETclass.BlankRows[1, 3:] + 2

    def readLEDETExcel(self, file):
        ##File must be whole eos string
        workbookVariables = openpyxl.load_workbook(file, data_only=True)

        #Inputs
        worksheetInputs = workbookVariables['Inputs']
        lastAttribute = worksheetInputs.cell(1, 2).value
        for i in range(1, worksheetInputs.max_row+1):
            self.variablesInputs[str(worksheetInputs.cell(i, 2).value)] = str(worksheetInputs.cell(i, 1).value)
            attribute = worksheetInputs.cell(i, 2).value
            try:
                if (attribute == None):
                    if worksheetInputs.cell(i, 3).value is not None:
                        values = read_row(worksheetInputs, i)
                        values = np.array([k for k in values if(str(k))])
                        current = self.getAttribute(self.Inputs, lastAttribute)
                        current = np.vstack((current, values))
                        self.setAttribute(self.Inputs, lastAttribute, current)
                    else:
                        continue
                elif (type(self.getAttribute(self.Inputs, attribute)) == np.ndarray):
                    lastAttribute = attribute
                    values = read_row(worksheetInputs, i)
                    values = np.array([k for k in values if(str(k))])
                    self.setAttribute(self.Inputs, attribute, values)
                else:
                    value = worksheetInputs.cell(i, 3).value
                    self.setAttribute(self.Inputs, attribute, value)
            except TypeError as e:
                if attribute in self.Inputs.__annotations__: raise e
                print("Error with attribute: {}, continuing.".format(attribute))
        #Options
        worksheetOptions = workbookVariables['Options']
        for i in range(1, worksheetOptions.max_row+1):
            self.variablesOptions[str(worksheetOptions.cell(i, 2).value)] = str(worksheetOptions.cell(i, 1).value)
            attribute = worksheetOptions.cell(i, 2).value
            try:
                if (type(self.getAttribute(self.Options, attribute)) == np.ndarray):
                    values = read_row(worksheetOptions, i)
                    values = np.array([k for k in values if(str(k))])
                    self.setAttribute(self.Options, attribute, values)
                else:
                    value = worksheetOptions.cell(i, 3).value
                    self.setAttribute(self.Options, attribute, value)
            except  TypeError as e:
                if attribute in self.Options.__annotations__: raise e
                print("Error with attribute: {}, continuing.".format(attribute))
        #Plots
        worksheetPlots = workbookVariables['Plots']
        for i in range(1, worksheetPlots.max_row+1):
            self.variablesPlots[str(worksheetPlots.cell(i, 2).value)] = str(worksheetPlots.cell(i, 1).value)
            attribute = worksheetPlots.cell(i, 2).value
            try:
                if (type(self.getAttribute(self.Plots, attribute)) == np.ndarray):
                    values = read_row(worksheetPlots, i, St=True)[2:]
                    values = np.array([k for k in values if(str(k))])
                    self.setAttribute(self.Plots, attribute, values)
                else:
                    try:
                        value = worksheetPlots.cell(i, 3).value
                    except:
                        value = ''
                    self.setAttribute(self.Plots, attribute, value)
            except  TypeError as e:
                print("Error with attribute: {}, continuing.".format(attribute))
        # Variables
        try:
            worksheetVariables = workbookVariables['Variables']
            for i in range(1, worksheetVariables.max_row+1):
                self.variablesVariables[str(worksheetVariables.cell(i, 2).value)] = str(worksheetVariables.cell(i, 1).value)
                attribute = worksheetVariables.cell(i, 2).value
                try:
                    if (type(self.getAttribute(self.Variables, attribute)) == np.ndarray):
                        if attribute != 'typeVariableToSaveTxt':  values = read_row(worksheetVariables, i, St = True)[2:]
                        else:  values = read_row(worksheetVariables, i)
                        values = np.array([k for k in values if(str(k))])
                        self.setAttribute(self.Variables, attribute, values)
                    else:
                        value = worksheetVariables.cell(i, 3).value
                        self.setAttribute(self.Variables, attribute, value)
                except TypeError as e:
                    if attribute in self.Variables.__annotations__: raise e
                    print("Error with attribute: {}, continuing.".format(attribute))
        except:
            pass
            print("Error while reading Variables. Please check!")

    def __init__(self):
        self.Inputs = LEDETInputs()
        self.Options = LEDETOptions()
        self.Plots = LEDETPlots()
        self.Variables = LEDETVariables()
        self.variablesInputs = {}
        self.variablesOptions = {}
        self.variablesPlots = {}
        self.variablesVariables = {}

        self.variableGroupInputs = asdict(self.Inputs)
        self.variableGroupOptions = asdict(self.Options)
        self.variableGroupPlots = asdict(self.Plots)
        self.variableGroupVariables = asdict(self.Variables)

        # Load and set the default LEDET parameters
        self.fileDefaultParameters = os.path.join('ledet', 'variableNamesDescriptions.xlsx')
        self.loadDefaultParameters(self.fileDefaultParameters)

    def setParameters(self, variablesInputs, variablesOptions, variablesPlots , variablesVariables):
        for k in variablesInputs.keys():
            self.setAttribute(LEDETInputs, k, variablesInputs[k])
        for k in variablesOptions.keys():
            self.setAttribute(LEDETInputs, k, variablesOptions[k])
        for k in variablesPlots.keys():
            self.setAttribute(LEDETInputs, k, variablesPlots[k])
        for k in variablesVariables.keys():
            self.setAttribute(LEDETInputs, k, variablesVariables[k])

        self.variablesInputs, self.variablesOptions, self.variablesPlots, self.variablesVariables = variablesInputs, variablesOptions, variablesPlots, variablesVariables

    def loadDefaultParameters(self, fileDefaultParameters: str):
        '''
            **Loads and sets the default LEDET parameters **

            Function to load and set the default LEDET parameters

            :param fileName: String defining the name of the file defining the default LEDET parameters
            :type fileName: str

            :return: None
        '''

        # Load default LEDET parameters
        # Read variable names and descriptions
        fullfileName = ResourceReader.getResourcePath(fileDefaultParameters)
        # print(fullfileName) # for debug
        workbookVariables = openpyxl.load_workbook(fullfileName)

        # Load "Inputs" sheet
        worksheetInputs = workbookVariables['Inputs']
        variablesInputs = {}
        for i in range(1, worksheetInputs.max_row+1):
            variablesInputs[str(worksheetInputs.cell(i, 2).value)] = str(worksheetInputs.cell(i, 1).value)

        # Load "Options" sheet
        worksheetOptions = workbookVariables['Options']
        variablesOptions = {}
        for i in range(1, worksheetOptions.max_row+1):
            variablesOptions[str(worksheetOptions.cell(i, 2).value)] = str(worksheetOptions.cell(i, 1).value)

        # Load "Plots" sheet
        worksheetPlots = workbookVariables['Plots']
        variablesPlots = {}
        for i in range(1, worksheetPlots.max_row+1):
            variablesPlots[str(worksheetPlots.cell(i, 2).value)] = str(worksheetPlots.cell(i, 1).value)

        # Load "Variables" sheet
        worksheetVariables = workbookVariables['Variables']
        variablesVariables = {}
        for i in range(1, worksheetVariables.max_row+1):
            variablesVariables[str(worksheetVariables.cell(i, 2).value)] = str(worksheetVariables.cell(i, 1).value)

        # Set parameters
        self.setParameters(variablesInputs, variablesOptions, variablesPlots, variablesVariables)

    def localsParser(self, locals):
        for attribute in locals:
            if attribute in self.Inputs.__annotations__:
                group = self.Inputs
            elif attribute in self.Options.__annotations__:
                group = self.Options
            elif attribute in self.Plots.__annotations__:
                group = self.Plots
            elif attribute in self.Variables.__annotations__:
                group = self.Variables
            else:
                continue
            tt = type(self.getAttribute(group, attribute))
            if tt == np.ndarray:
                    self.setAttribute(group, attribute, np.array(locals[attribute]))
            else:
                self.setAttribute(group, attribute, locals[attribute])

    def cpCu_nist_mat(self, T):
        density = 8960
        cpCu_perMass = np.zeros(T.size)
        T[T < 4] = 4
        idxT1 = np.where(T < 300)
        idxT2 = np.where(T >= 300)
        dc_a = -1.91844
        dc_b = -0.15973
        dc_c = 8.61013
        dc_d = -18.996
        dc_e = 21.9661
        dc_f = -12.7328
        dc_g = 3.54322
        dc_h = -0.3797

        logT1 = np.log10(T[idxT1])
        tempVar = \
        dc_a + dc_b * (logT1)**1 + dc_c * (logT1)**2 + dc_d * (logT1)**3 + \
        dc_e * (logT1)**4 + dc_f * (logT1)**5 + dc_g * (logT1)** 6 + dc_h * (logT1)**7
        cpCu_perMass[idxT1] = 10**tempVar

        cpCu_perMass[idxT2]= 361.5 + 0.093 * T[idxT2]
        cpCu = density * cpCu_perMass
        return cpCu

    def cpNbTi_cudi_mat(self, T, B):
        Tc0 = 9.2
        Bc20 = 14.5
        alpha = .59
        B[B>= Bc20] = Bc20-10E-4

        Tc = Tc0 * (1 - B / Bc20)**alpha
        cpNbTi = np.zeros(T.size)

        idxT1 = np.where(T <= Tc)
        idxT2 = np.where((T > Tc) & (T <= 20.0))
        idxT3 = np.where((T > 20) & (T <= 50))
        idxT4 = np.where((T > 50) & (T <= 175))
        idxT5 = np.where((T > 175) & (T <= 500))
        idxT6 = np.where((T > 500) & (T <= 1000))
        idxT7 = np.where(T > 1000)

        p1 = [0.00000E+00,    4.91000E+01,   0.00000E+00,   6.40000E+01,  0.00000E+00]
        p2 = [0.00000E+00,   1.62400E+01,   0.00000E+00,  9.28000E+02,   0.00000E+00]
        p3 = [-2.17700E-01,   1.19838E+01,   5.53710E+02, - 7.84610E+03,  4.13830E+04]
        p4 = [-4.82000E-03,  2.97600E+00, -7.16300E+02,  8.30220E+04,  -1.53000E+06]
        p5 = [-6.29000E-05, 9.29600E-02, -5.16600E+01,  1.37060E+04,  1.24000E+06]
        p6 = [0.00000E+00, 0.00000E+00,  -2.57000E-01,  9.55500E+02,  2.45000E+06]
        p7 = [0, 0, 0, 0, 3.14850E+06]

        cpNbTi[idxT1] = p1[0] * T[idxT1]**4 + p1[1] * T[idxT1]**3 + p1[2] * T[idxT1]**2 + p1[3] * T[idxT1] + p1[4]
        cpNbTi[idxT2] = p2[0] * T[idxT2]**4 + p2[1] * T[idxT2]**3 + p2[2] * T[idxT2]**2 + p2[3] * T[idxT2] + p2[4]
        cpNbTi[idxT3] = p3[0] * T[idxT3]**4 + p3[1] * T[idxT3]**3 + p3[2] * T[idxT3]**2 + p3[3] * T[idxT3] + p3[4]
        cpNbTi[idxT4] = p4[0] * T[idxT4]**4 + p4[1] * T[idxT4]**3 + p4[2] * T[idxT4]**2 + p4[3] * T[idxT4] + p4[4]
        cpNbTi[idxT5] = p5[0] * T[idxT5]**4 + p5[1] * T[idxT5]**3 + p5[2] * T[idxT5]**2 + p5[3] * T[idxT5] + p5[4]
        cpNbTi[idxT6] = p6[0] * T[idxT6]**4 + p6[1] * T[idxT6]**3 + p6[2] * T[idxT6]**2 + p6[3] * T[idxT6] + p6[4]
        cpNbTi[idxT7] = p7[0] * T[idxT7]**4 + p7[1] * T[idxT7]**3 + p7[2] * T[idxT7]**2 + p7[3] * T[idxT7] + p7[4]
        return cpNbTi

    def quenchPropagationVelocity(self, I, B, T_bath, cable):
        L0 = 2.44E-08
        A_CableBare = cable.A_CableInsulated * (cable.f_SC + cable.f_ST)
        f_SC_inStrand = cable.f_SC / (cable.f_SC + cable.f_ST)
        f_ST_inStrand = cable.f_ST / (cable.f_SC + cable.f_ST)
        I = abs(I)
        J_op = I / A_CableBare
        Tc = cable.Tc0_NbTi * (1 - B / cable.Bc20_NbTi) ** cable.alpha_NbTi
        Tcs= (1 - I / (cable.c1_Ic_NbTi + cable.c2_Ic_NbTi * B)) * Tc
        Ts = (Tcs + Tc) / 2
        cp_ST = self.cpCu_nist_mat(Ts)
        cp_SC = self.cpNbTi_cudi_mat(Ts, B)
        cp = cp_ST * f_ST_inStrand + cp_SC * f_SC_inStrand
        vQ = J_op / cp * ((L0 * Ts) / (Ts - T_bath))**0.5
        idxInfQuenchVelocity=np.where(Tcs <= T_bath)
        vQ[idxInfQuenchVelocity]=1E6
        return vQ

    def acquireBField(self, ROXIE_File):
        reader = csv.reader(open(ROXIE_File), delimiter="\t")
        Inom = self.Options.Iref
        reader = csv.reader(open(ROXIE_File))
        B_Field = np.array([])
        stack = 0
        for row in reader:
            if not row: continue
            row_s = np.array(row[0].split())
            if not stack:
                B_Field = np.array(row_s[1:])
                stack = 1
            else:
                B_Field = np.vstack((B_Field, np.array(row_s)))
        B_Field = B_Field[1:].astype(float)
        BX = B_Field[:, 5].transpose()
        BY = B_Field[:, 6].transpose()
        f_mag_X_all = BX / Inom
        f_mag_Y_all = BY / Inom
        f_mag = (f_mag_X_all**2 + f_mag_Y_all**2) ** 0.5
        B = f_mag * self.Inputs.I00

        B[B > 10E6]=10E-6
        return B

    def repeatCable(self, cable):
        nT = self.Inputs.nT
        nT = nT.astype(int)
        newCable = Cable()
        for attribute in cable.__annotations__:
            if attribute == 'alpha_NbTi': continue
            x = np.ndarray([])
            x = getattr(cable, attribute)
            x = np.repeat(x, nT)
            setattr(newCable, attribute, x)
        return newCable

    def adjust_vQ(self, ROXIE_File):
        cable = Cable()
        cable.A_CableInsulated = (self.Inputs.wBare_inGroup+2*self.Inputs.wIns_inGroup) \
                               * (self.Inputs.hBare_inGroup+2*self.Inputs.hIns_inGroup)
        B = self.acquireBField(ROXIE_File)
        if max(self.Inputs.nStrands_inGroup) > 1:
            strandCount = 0
            GroupCount = 0
            Bcopy = np.zeros(int(sum(self.Inputs.nT)))
            for i in range(int(sum(self.Inputs.nT))):
                Bcopy[i] = sum(B[int(strandCount):int(strandCount+self.Inputs.nStrands_inGroup[GroupCount])])/self.Inputs.nStrands_inGroup[int(GroupCount)]
                TurnSum = sum(self.Inputs.nT[0:GroupCount+1])
                if i>TurnSum: GroupCount = GroupCount + 1
                strandCount = strandCount + self.Inputs.nStrands_inGroup[GroupCount]
            B = Bcopy
            cable.f_SC = self.Inputs.f_SC_strand_inGroup * \
                         (self.Inputs.nStrands_inGroup* (np.pi/4)*(self.Inputs.ds_inGroup**2)) / cable.A_CableInsulated
            cable.f_ST = (1 - self.Inputs.f_SC_strand_inGroup) * \
                         (self.Inputs.nStrands_inGroup* (np.pi/4)*(self.Inputs.ds_inGroup**2)) / cable.A_CableInsulated
        else:
            cable.f_SC = self.Inputs.f_SC_strand_inGroup * \
                         (self.Inputs.wBare_inGroup * self.Inputs.hBare_inGroup) / cable.A_CableInsulated
            cable.f_ST = (1 - self.Inputs.f_SC_strand_inGroup) * \
                         (self.Inputs.wBare_inGroup * self.Inputs.hBare_inGroup) / cable.A_CableInsulated

        I = self.Inputs.I00
        T_bath = self.Inputs.T00
        cable.SCtype = self.Inputs.SCtype_inGroup
        cable.STtype = self.Inputs.STtype_inGroup
        cable.Tc0_NbTi = self.Inputs.Tc0_NbTi_ht_inGroup
        cable.Bc20_NbTi = self.Inputs.Bc2_NbTi_ht_inGroup
        cable.c1_Ic_NbTi = self.Inputs.c1_Ic_NbTi_inGroup
        cable.c2_Ic_NbTi = self.Inputs.c2_Ic_NbTi_inGroup
        cable.alpha_NbTi = .59
        cable.Jc_Nb3Sn0 = self.Inputs.Jc_Nb3Sn0_inGroup
        cable.Tc0_Nb3Sn = self.Inputs.Tc0_Nb3Sn_inGroup
        cable.Bc20_Nb3Sn = self.Inputs.Bc2_Nb3Sn_inGroup
        cable = self.repeatCable(cable)

        vQ = self.quenchPropagationVelocity(I, B, T_bath, cable)
        self.setAttribute(getattr(self, "Inputs"), "vQ_iStartQuench", vQ)
        return #vQ

    def addVariablesInputs(self,
                           T00, l_magnet, I00, M_m,
                           fL_I, fL_L,
                           GroupToCoilSection, polarities_inGroup, nT, nStrands_inGroup, l_mag_inGroup, ds_inGroup,
                           f_SC_strand_inGroup, f_ro_eff_inGroup, Lp_f_inGroup, RRR_Cu_inGroup,
                           SCtype_inGroup, STtype_inGroup, insulationType_inGroup, internalVoidsType_inGroup,
                           externalVoidsType_inGroup,
                           wBare_inGroup, hBare_inGroup, wIns_inGroup, hIns_inGroup, Lp_s_inGroup, R_c_inGroup,
                           Tc0_NbTi_ht_inGroup, Bc2_NbTi_ht_inGroup, c1_Ic_NbTi_inGroup, c2_Ic_NbTi_inGroup,
                           Tc0_Nb3Sn_inGroup, Bc2_Nb3Sn_inGroup, Jc_Nb3Sn0_inGroup,
                           el_order_half_turns,
                           alphasDEG, rotation_block, mirror_block, mirrorY_block,
                           iContactAlongWidth_From, iContactAlongWidth_To, iContactAlongHeight_From,
                           iContactAlongHeight_To,
                           iStartQuench, tStartQuench, lengthHotSpot_iStartQuench, vQ_iStartQuench,
                           R_circuit, R_crowbar, Ud_crowbar, t_PC, t_PC_LUT, I_PC_LUT,
                           tEE, R_EE_triggered,
                           tCLIQ, directionCurrentCLIQ, nCLIQ, U0, C, Rcapa,
                           tQH, U0_QH, C_QH, R_warm_QH, w_QH, h_QH, s_ins_QH, type_ins_QH, s_ins_QH_He, type_ins_QH_He, l_QH, f_QH,
                           iQH_toHalfTurn_From, iQH_toHalfTurn_To,
                           tQuench, initialQuenchTemp,
                           HalfTurnToInductanceBlock, M_InductanceBlock_m
                           ):
        '''
            **Adds all LEDET parameters to be written in the "Inputs" sheet **

            Function to add "Inputs" LEDET parameters

            :param T00: String defining the name of the file defining the default LEDET parameters
            :type T00: float

            :return: None
        '''
        ins = locals()
        for attribute in ins:
            self.setAttribute(self.Inputs, attribute, ins[attribute])

    def addVariablesOptions(self,
                            time_vector_params,
                            Iref, flagIron, flagSelfField, headerLines, columnsXY, columnsBxBy, flagPlotMTF,
                            flag_calculateInductanceMatrix, flag_useExternalInitialization, flag_initializeVar,
                            flag_fastMode, flag_controlCurrent, flag_automaticRefinedTimeStepping, flag_IronSaturation,
                            flag_InvertCurrentsAndFields, flag_ScaleDownSuperposedMagneticField, flag_HeCooling,
                            fScaling_Pex, fScaling_Pex_AlongHeight,
                            fScaling_MR, flag_scaleCoilResistance_StrandTwistPitch, flag_separateInsulationHeatCapacity,
                            flag_ISCL, fScaling_Mif, fScaling_Mis, flag_StopIFCCsAfterQuench, flag_StopISCCsAfterQuench,
                            tau_increaseRif, tau_increaseRis,
                            fScaling_RhoSS, maxVoltagePC, flag_symmetricGroundingEE, flag_removeUc, BtX_background,
                            BtY_background,
                            flag_showFigures, flag_saveFigures, flag_saveMatFile, flag_saveTxtFiles,
                            flag_generateReport,
                            flag_hotSpotTemperatureInEachGroup, MinMaxXY_MTF
                           ):
        '''
            **Adds all LEDET parameters to be written in the "Options" sheet **

            Function to add "Options" LEDET parameters

            :param T00: String defining the name of the file defining the default LEDET parameters
            :type T00: float

            :return: None
        '''
        ins = locals()
        for attribute in ins:
            self.setAttribute(self.Options, attribute, ins[attribute])

    def addVariablesPlots(self,
                          suffixPlot, typePlot, outputPlotSubfolderPlot, variableToPlotPlot, selectedStrandsPlot,
                          selectedTimesPlot,
                          labelColorBarPlot, minColorBarPlot, maxColorBarPlot, MinMaxXYPlot, flagSavePlot,
                          flagColorPlot, flagInvisiblePlot
                           ):
        '''
            **Adds all LEDET parameters to be written in the "Plots" sheet **

            Function to add "Plots" LEDET parameters

            :param T00: String defining the name of the file defining the default LEDET parameters
            :type T00: float

            :return: None
        '''
        ins = locals()
        for attribute in ins:
            self.setAttribute(self.Plots, attribute, ins[attribute])

    def addVariablesVariables(self,
                              variableToSaveTxt, typeVariableToSaveTxt, variableToInitialize
                              ):
        '''
            **Adds all LEDET parameters to be written in the "Variables" sheet **

            Function to add "Variables" LEDET parameters

            :param T00: String defining the name of the file defining the default LEDET parameters
            :type T00: float

            :return: None
        '''
        ins = locals()
        for attribute in ins:
            self.setAttribute(self.Variables, attribute, ins[attribute])


    def printVariableDescNameValue(self, variableGroup, variableLabels):
        """

           **Print variable description, variable name, and variable value**

           Function prints variable description, variable name, and variable value

           :param variableGroup: list of tuples; each tuple has two elements: the first element is a string defining
           the variable name, and the second element is either an integer, a float, a list, or a numpy.ndarray
           defining the variable value :type variableGroup: list :param variableLabels: dictionary assigning a
           description to each variable name
           :type variableLabels: dict

           :return: None

           - Example :

           import numpy as np

            variableGroup = []
            variableGroup.append( ('x1', 12) )
            variableGroup.append( ('x2', 23.42) )
            variableGroup.append( ('x3', [2, 4, 6]) )
            variableGroup.append( ('x3', np.array([2, 4, 6])) )

            variableLabels = {'x1': '1st variable', 'x2': '2nd variable', 'x3': '3rd variable'}

            utils.printVariableDescNameValue(variableGroup, variableLabels)
            # >>> 					1st variable x1 12
            # >>> 					2nd variable x2 23.42
            # >>> 					3rd variable x3 [2, 4, 6]
            # >>> 					3rd variable x3 [2 4 6]

        """
        if(variableGroup == self.variableGroupInputs):
            variableGroup=self.Inputs
        if (variableGroup == self.variableGroupOptions):
            variableGroup = self.Options
        if (variableGroup == self.variableGroupPlots):
            variableGroup = self.Plots
        if (variableGroup == self.variableGroupVariables):
            variableGroup = self.Variables

        if(type(variableGroup) != dict):
            for k in variableGroup.__annotations__:
                if (k == "BlankRows"): continue
                if k == 'overwrite_f_internalVoids_inGroup': continue
                if k == 'overwrite_f_externalVoids_inGroup': continue
                print(variableLabels[k])
                print(k, self.getAttribute(variableGroup, k))
        else:
            for k in variableGroup:
                if (k == "BlankRows"): continue
                if k =='overwrite_f_internalVoids_inGroup': continue
                if k == 'overwrite_f_externalVoids_inGroup': continue
                print(variableLabels[k])
                print(k, variableGroup[k])

    def getNumberOfCoilSections(self):
        k = self.Inputs.M_m
        if k.shape == (1,): return k.shape[0]
        try:
            if k.shape[0] != k.shape[1]: print("M_m is not square")
        except:
            print("M_m is not square")
            return -1
        return k.shape[0]

    def checkM_InductanceBlock_m(self, Turns):
        if type(self.Inputs.M_InductanceBlock_m) != np.ndarray:
            k = np.array(self.Inputs.M_InductanceBlock_m)
        else:
            k = self.Inputs.M_InductanceBlock_m
        if k.shape == (1,):
            return False
        try:
            if k.shape[0] == k.shape[1]:
                if k.shape[0] == Turns:
                    return False
        except:
            print("M_InductanceBlock_m is not correct!")
            return True
        print("M_InductanceBlock_m is not correct!")
        return True

    def checkHeFraction(self, Groups):
        k = self.Inputs.overwrite_f_externalVoids_inGroup
        k2 = self.Inputs.overwrite_f_internalVoids_inGroup
        if len(k) > 0:
            if len(k) != len(k2):
                print("Helium section was set but is corrupted.")
                return False
            if len(k) != Groups:
                print("Helium section was set but is wrong length.")
                return False
        elif len(k2) > 0:
            print("Helium section was set but is corrupted.")
            return False
        return True

    def checkMonotony(self, arr, Invert=False):
        if Invert:
            arr = np.flip(arr)
        b = all(x <= y for x, y in zip(arr, arr[1:]))
        return b

    def checkTimes(self):
        t1 = min(self.Inputs.tQuench)
        t2 = self.Inputs.t_PC_LUT[0]
        t3 = min(self.Inputs.tStartQuench)
        t4 = self.Options.time_vector_params[0]

        if any(x < t4 for x in [t1,t2,t3]):
            print("You're using a time, that is before the start of the simulation. Please check!")
            return 0
        else:
            return 1

    def consistencyCheckLEDET(self):
        ## Consistency check of Inputs
        ## 0 Single - 1 CoilSections - 2 Groups - 3 Half-Turns - 4 doesn't matter - 5 iContactAlongWidth - 6 iContactAlongHeight - 7 vQlength
        ## 8 Quench Heater, 9 QuenchToFrom, 10 CLIQ,
        slicesSameInput = [[0,1,2,49,50,51,52,55,56],
                           [3,58,77,78],
                           [6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33],
                           [36,37,38,39,40,79],
                           [4,5,34,35,53,54],
                           [41,42],
                           [43,44],
                           [45,46,47,48],
                           [63,64,65,66,67,68,69,70,71,72,73,74],
                           [75,76],
                           [57,59,60,61,62]]
        lengthInputs = len(self.Inputs.__annotations__)
        sizeInputs = np.zeros((lengthInputs,1))
        sizeInputs[slicesSameInput[0]] = 1
        sizeInputs[slicesSameInput[1]] = self.getNumberOfCoilSections()
        if sizeInputs[slicesSameInput[1][0]] == -1:
            print("M_m is corrupted. please check.")
            return True
        sizeInputs[slicesSameInput[2]] = len(self.Inputs.nT)
        sizeInputs[slicesSameInput[3]] = sum(self.Inputs.nT)
        sizeInputs[slicesSameInput[4]] = 0
        sizeInputs[slicesSameInput[5]] = len(self.Inputs.iContactAlongWidth_From)
        sizeInputs[slicesSameInput[6]] = len(self.Inputs.iContactAlongHeight_From)
        sizeInputs[slicesSameInput[7]] = len(self.Inputs.iStartQuench)
        sizeInputs[slicesSameInput[8]] = len(self.Inputs.tQH)
        sizeInputs[slicesSameInput[9]] = len(self.Inputs.iQH_toHalfTurn_From)
        try:
            sizeInputs[slicesSameInput[10]] = len(self.Inputs.tCLIQ)
        except:
            sizeInputs[slicesSameInput[10]] = 1

        if self.checkM_InductanceBlock_m(int(sum(self.Inputs.nT)/2)):
            return 1

        Count = 0
        Break = 0
        for k in self.Inputs.__annotations__:
            if k== "BlankRows": continue
            if sizeInputs[Count] == 0:
                Count = Count + 1
                continue
            cC = self.getAttribute(self.Inputs, k)
            if type(cC) == np.ndarray or type(cC) == list:
                if not len(cC)==sizeInputs[Count]:
                    print("The variable ",k, " does not have the correct size, should be", sizeInputs[Count]," but is ",len(cC),"! Please check.")
                    Break = 1
            elif type(cC) == float or type(cC) == int or type(cC) == np.float64:
                if not sizeInputs[Count]==1:
                    print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is ",len(cC)," Please check.")
                    Break = 1
            else:
                print("Variable ", k, " has the wrong data-type set! Please check.")
                Break = 1
            Count = Count + 1
        if not self.checkHeFraction(len(self.Inputs.nT)): Break = 1
        if not self.checkMonotony(self.Inputs.t_PC_LUT):
            print("t_PC_LUT is not monotonic")
            Break = 1
        # if not self.checkMonotony(self.Inputs.fL_L, Invert=True):
        #     print("fL_L is not monotonic")
        #     Break = 1
        if not self.checkMonotony(self.Inputs.fL_I):
            print("fL_I is not monotonic")
            Break = 1
        if not self.checkTimes():
            Break = 1

        return Break

    def writeFileLEDET(self, nameFileLEDET, verbose: bool = False, SkipConsistencyCheck: bool = False):
        '''
            **Writes LEDET input file **

            Function to write a LEDET input file composed of "Inputs", "Options", "Plots", and "Variables" sheets

            :param nameFileLEDET: String defining the name of the LEDET input file to be written
            :type nameFileLEDET: string
            :param verbose: flag that determines whether the output are printed
            :type verbose: bool

            :return: None
        '''
        if not SkipConsistencyCheck:
            if self.consistencyCheckLEDET():
                print("Variables are not consistent! Writing aborted - ", nameFileLEDET)
                return
            else:
                print("Preliminary consistency check was successful! - ", nameFileLEDET)

        workbook = openpyxl.Workbook()

        if verbose:
            print('')
            print('### Write "Variables" sheet ###')
        self.writeLEDETInputsNew(workbook, "Variables", self.Variables, self.variablesVariables, verbose)

        if verbose:
            print('')
            print('### Write "Plots" sheet ###')
        self.writeLEDETInputsNew(workbook, "Plots", self.Plots, self.variablesPlots, verbose)

        if verbose:
            print('')
            print('### Write "Options" sheet ###')
        self.writeLEDETInputsNew(workbook, "Options", self.Options, self.variablesOptions, verbose)

        if verbose:
            print('### Write "Inputs" sheet ###')
        self.writeLEDETInputsNew(workbook, "Inputs", self.Inputs, self.variablesInputs, verbose)

        # Save the workbook
        std = workbook['Sheet']
        workbook.remove(std)
        workbook.save(nameFileLEDET)

        # Display time stamp and end run
        currentDT = datetime.datetime.now()
        if verbose:
            print(' ')
            print('Time stamp: ' + str(currentDT))
            print('New file ' + nameFileLEDET + ' generated.')

        return

    def writeLEDETInputsNew(self, book, sheet, variableGroup, variableLabels, verbose: bool = False):
        """
            **Write one sheet of a LEDET input file**

            Function writes one sheet of a LEDET input file

            :param book: workbook object to write
            :type book: xlsxwriter.Workbook
            :param sheet: name of the sheet to write (first sheet = 0)
            :type sheet: string
            :param variableGroup: list of tuples; each tuple has two elements: the first element is a string defining the variable name, and the second element is either an integer, a float, a list, or a numpy.ndarray defining the variable value
            :param variableLabels: dictionary assigning a description to each variable name
            :type variableLabels: dict
            :param verbose: flag that determines whether the output are printed
            :type verbose: bool
            :return:
        """
        book.create_sheet(index = 1 , title = sheet)
        for s in range(len(book.sheetnames)):
            if book.sheetnames[s] == sheet: break
        book.active = s
        sheet1 = book.active

        # Write to the sheet of the workbook
        currentRow = 0
        self.updateBlankrows(variableGroup)
        for attribute in variableGroup.__annotations__:
            if (attribute == "BlankRows"): continue
            if (attribute == "overwrite_f_internalVoids_inGroup"):
                ofiVg = self.getAttribute(variableGroup, attribute)
                if len(ofiVg) == 0: continue
            if (attribute == "overwrite_f_externalVoids_inGroup"):
                ofiVg = self.getAttribute(variableGroup, attribute)
                if len(ofiVg) == 0: continue
            if currentRow in variableGroup.BlankRows[1]:
                currentRow = currentRow + 1
                sheet1.append([None])

            varDesc = variableLabels.get(str(attribute))
            varType = type(self.getAttribute(variableGroup, attribute))
            if varType == np.ndarray:
                if self.getAttribute(variableGroup, attribute).ndim > 1:
                    for i in range(self.getAttribute(variableGroup, attribute).shape[1]):
                        values = self.getAttribute(variableGroup, attribute)[i].tolist()
                        if i == 0 : sheet1.append([varDesc, attribute]+ values)
                        else : sheet1.append([None, None]+ values)
                        currentRow = currentRow + 1
                    continue

            currentRow = currentRow + 1
            if isinstance(self.getAttribute(variableGroup, attribute), np.ndarray):
                values = np.array(self.getAttribute(variableGroup, attribute)).tolist()
            elif isinstance(self.getAttribute(variableGroup, attribute), list):
                values = self.getAttribute(variableGroup, attribute)
            else: values = [self.getAttribute(variableGroup, attribute)]
            sheet1.append([varDesc, attribute]+ values)

            width = [80.7109375, 40.7109375, 20.7109375]
            sheet1.column_dimensions['A'].width = width[0]
            sheet1.column_dimensions['B'].width = width[1]
            maxc = get_column_letter(sheet1.max_column)
            for i in range(3,sheet1.max_column+1):
                cl = get_column_letter(i)
                sheet1.column_dimensions[cl].width = width[2]